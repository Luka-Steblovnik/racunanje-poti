"""
Kilometer Tracker — FastAPI backend with authentication
"""

import os
import io
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path

import bcrypt
import httpx
import jwt
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel

load_dotenv()

GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-please-change")
TOKEN_EXPIRE_DAYS = 30
ALGORITHM = "HS256"

_data_dir = Path(os.getenv("DATA_DIR", str(Path(__file__).parent)))
DB_FILE = _data_dir / "app.db"

app = FastAPI(title="Kilometer Tracker")

_raw_origins = os.getenv("ALLOWED_ORIGINS", "http://localhost:5173,http://localhost:3000")
_origins = [o.strip() for o in _raw_origins.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_origin_regex=r"https://.*\.pages\.dev" if "*" not in _origins else None,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Database ──────────────────────────────────────────────────────────────────

def get_db():
    _data_dir.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_FILE))
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at    TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS routes (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            datetime    TEXT NOT NULL,
            origin      TEXT NOT NULL,
            destination TEXT NOT NULL,
            distance_km REAL NOT NULL,
            duration    TEXT NOT NULL,
            source      TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    conn.commit()
    conn.close()


init_db()


# ── Auth helpers ──────────────────────────────────────────────────────────────

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())


def create_token(user_id: int, username: str) -> str:
    payload = {
        "sub": str(user_id),
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(days=TOKEN_EXPIRE_DAYS),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


security = HTTPBearer()


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        return {"id": int(payload["sub"]), "username": payload["username"]}
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Seja je potekla — prosim prijavi se znova")
    except Exception:
        raise HTTPException(401, "Neveljaven token")


# ── Routing helpers ───────────────────────────────────────────────────────────

def fmt_duration(seconds):
    minutes = seconds // 60
    h, m = divmod(minutes, 60)
    if h > 0:
        return f"{h} h {m} min"
    return f"{m} min"


async def geocode_nominatim(address):
    params = {"q": address, "format": "json", "limit": 1}
    headers = {"User-Agent": "KilometerTracker/1.0"}
    async with httpx.AsyncClient(timeout=10) as client:
        r = await client.get("https://nominatim.openstreetmap.org/search", params=params, headers=headers)
    results = r.json()
    if not results:
        raise ValueError(f"Naslov '{address}' ni bil najden.")
    return float(results[0]["lat"]), float(results[0]["lon"])


async def calc_osrm(origin, destination, origin_lat=None, origin_lon=None, dest_lat=None, dest_lon=None):
    lat1, lon1 = (origin_lat, origin_lon) if origin_lat is not None else await geocode_nominatim(origin)
    lat2, lon2 = (dest_lat, dest_lon) if dest_lat is not None else await geocode_nominatim(destination)
    url = f"http://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}"
    async with httpx.AsyncClient(timeout=15) as client:
        r = await client.get(url, params={"overview": "false"})
    data = r.json()
    if data.get("code") != "Ok":
        raise ValueError("OSRM: pot ni bila najdena.")
    route = data["routes"][0]
    return round(route["distance"] / 1000, 1), fmt_duration(int(route["duration"]))


# ── Pydantic models ───────────────────────────────────────────────────────────

class AuthRequest(BaseModel):
    username: str
    password: str


class CalculateRequest(BaseModel):
    origin: str
    destination: str
    origin_lat: float = None
    origin_lon: float = None
    dest_lat: float = None
    dest_lon: float = None


class SaveRequest(BaseModel):
    origin: str
    destination: str
    distance_km: float
    duration: str
    source: str


# ── Auth endpoints ────────────────────────────────────────────────────────────

@app.post("/auth/register")
async def register(req: AuthRequest):
    if len(req.username.strip()) < 3:
        raise HTTPException(400, "Uporabniško ime mora imeti vsaj 3 znake")
    if len(req.password) < 6:
        raise HTTPException(400, "Geslo mora imeti vsaj 6 znakov")
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
            (req.username.strip(), hash_password(req.password), datetime.now().isoformat()),
        )
        conn.commit()
        user_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        return {"token": create_token(user_id, req.username.strip()), "username": req.username.strip()}
    except sqlite3.IntegrityError:
        raise HTTPException(400, "To uporabniško ime je že zasedeno")
    finally:
        conn.close()


@app.post("/auth/login")
async def login(req: AuthRequest):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id, username, password_hash FROM users WHERE username = ?",
            (req.username.strip(),),
        ).fetchone()
        if not row or not verify_password(req.password, row["password_hash"]):
            raise HTTPException(401, "Napačno uporabniško ime ali geslo")
        return {"token": create_token(row["id"], row["username"]), "username": row["username"]}
    finally:
        conn.close()


# ── Protected endpoints ───────────────────────────────────────────────────────

@app.post("/calculate")
async def calculate(req: CalculateRequest, user=Depends(get_current_user)):
    if not req.origin.strip() or not req.destination.strip():
        raise HTTPException(400, "Izhodišče in cilj ne smeta biti prazna.")
    try:
        km, dur = await calc_osrm(
            req.origin.strip(), req.destination.strip(),
            req.origin_lat, req.origin_lon, req.dest_lat, req.dest_lon,
        )
        maps_url = (
            "https://www.google.com/maps/dir/?api=1"
            f"&origin={req.origin.replace(' ', '+')}"
            f"&destination={req.destination.replace(' ', '+')}"
        )
        return {"distance_km": km, "duration": dur, "maps_url": maps_url, "source": "osrm"}
    except ValueError as e:
        raise HTTPException(400, str(e))
    except httpx.TimeoutException:
        raise HTTPException(503, "Zahteva je potekla. Poskusi znova.")
    except Exception as e:
        raise HTTPException(500, f"Nepričakovana napaka: {e}")


@app.post("/routes")
async def save_route(req: SaveRequest, user=Depends(get_current_user)):
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO routes (user_id, datetime, origin, destination, distance_km, duration, source) VALUES (?,?,?,?,?,?,?)",
            (user["id"], datetime.now().isoformat(timespec="seconds"),
             req.origin, req.destination, req.distance_km, req.duration, req.source),
        )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.get("/routes")
async def get_routes(user=Depends(get_current_user)):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM routes WHERE user_id = ? ORDER BY datetime ASC", (user["id"],)
        ).fetchall()
        routes = [dict(r) for r in rows]
        return {"routes": routes, "total_km": round(sum(r["distance_km"] for r in routes), 1)}
    finally:
        conn.close()


@app.get("/routes/export")
async def export_xlsx(user=Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM routes WHERE user_id = ? ORDER BY datetime ASC", (user["id"],)
        ).fetchall()
        routes = [dict(r) for r in rows]
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prevozeni kilometri"

    ws.append(["Datum", "Od", "Kam", "Kilometri"])
    fill = PatternFill("solid", fgColor="2563EB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for r in routes:
        ws.append([r["datetime"][:10], r["origin"], r["destination"], r["distance_km"]])

    if routes:
        total = round(sum(r["distance_km"] for r in routes), 1)
        ws.append([])
        ws.append(["SKUPAJ", "", "", total])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 13

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prevozeni_kilometri.xlsx"},
    )


# ── Health check ─────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"ok": True}


# ── Autocomplete (no auth required) ──────────────────────────────────────────

@app.get("/autocomplete")
async def autocomplete(q: str = ""):
    if not q or len(q.strip()) < 2:
        return {"suggestions": []}
    params = {
        "q": q, "format": "json", "limit": 7,
        "addressdetails": 1, "countrycodes": "si,hr,at,it",
        "accept-language": "sl,en",
    }
    try:
        async with httpx.AsyncClient(timeout=8) as client:
            r = await client.get(
                "https://nominatim.openstreetmap.org/search",
                params=params, headers={"User-Agent": "KilometerTracker/1.0"},
            )
        results = r.json()
        suggestions = []
        seen = set()
        for item in results:
            a = item.get("address", {})
            road = a.get("road") or a.get("pedestrian") or a.get("path") or ""
            hnum = a.get("house_number") or ""
            city = a.get("city") or a.get("town") or a.get("village") or a.get("municipality") or ""
            postcode = a.get("postcode") or ""
            country = a.get("country") or ""
            lat, lon = float(item["lat"]), float(item["lon"])
            if road:
                main = f"{road} {hnum}".strip() if hnum else road
                if city:
                    main += f", {city}"
            elif city:
                main = city
            else:
                main = item.get("display_name", "").split(",")[0].strip()
            sub = ", ".join(s for s in [postcode, country] if s)
            key = main.lower()
            if key in seen:
                continue
            seen.add(key)
            suggestions.append({"id": f"{lat},{lon}", "main": main, "sub": sub, "lat": lat, "lon": lon})
            if len(suggestions) >= 5:
                break
        return {"suggestions": suggestions}
    except Exception:
        return {"suggestions": []}
